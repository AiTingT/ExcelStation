from pathlib import Path
from typing import Iterator, List, Tuple, Optional
import csv
import logging
import io

logger = logging.getLogger(__name__)

try:
    from python_calamine import CalamineWorkbook
    HAS_CALAMINE = True
except ImportError:
    HAS_CALAMINE = False
    logger.warning("python-calamine 未安装，将使用 openpyxl 作为后备")

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import msoffcrypto
    HAS_MSOFFCRYPTO = True
except ImportError:
    HAS_MSOFFCRYPTO = False

try:
    import xlrd
    HAS_XLRD = True
except ImportError:
    HAS_XLRD = False

try:
    import polars as pl
    HAS_POLARS = True
except ImportError:
    HAS_POLARS = False
    logger.warning("polars 未安装，将使用 python-calamine/openpyxl 作为后备")


class FileEncryptedError(Exception):
    """文件被加密的异常"""
    def __init__(self, message="文件已加密，需要密码才能打开"):
        self.message = message
        super().__init__(self.message)


class ExcelParserService:
    """Excel 文件解析服务 - 多重回退机制，支持加密文件"""

    @staticmethod
    def is_file_encrypted(file_path: Path) -> bool:
        """检测文件是否加密"""
        suffix = file_path.suffix.lower()

        if suffix == ".xlsx":
            try:
                with open(file_path, "rb") as f:
                    header = f.read(8)
                    if header.startswith(b"PK\x03\x04"):
                        return False
                    if header.startswith(b"\xd0\xcf\x11\xe0"):
                        try:
                            import msoffcrypto
                            f.seek(0)
                            office_file = msoffcrypto.OfficeFile(f)
                            return office_file.is_encrypted()
                        except Exception:
                            return True
            except Exception:
                pass
            return False

        elif suffix == ".xls":
            try:
                import xlrd
                try:
                    xlrd.open_workbook(str(file_path))
                    return False
                except xlrd.biffh.XLRDError as e:
                    if "encrypted" in str(e).lower() or "password" in str(e).lower():
                        return True
                    raise
            except Exception:
                pass
            return False

        return False

    @staticmethod
    def decrypt_file_to_memory(file_path: Path, password: str = "VelvetSweatshop") -> Optional[io.BytesIO]:
        """
        尝试解密文件到内存
        默认密码 VelvetSweatshop 是 Excel 的默认保护密码（很多企业加密用这个）
        """
        if not HAS_MSOFFCRYPTO:
            return None

        try:
            with open(file_path, "rb") as f:
                office_file = msoffcrypto.OfficeFile(f)
                if not office_file.is_encrypted():
                    f.seek(0)
                    return io.BytesIO(f.read())

                try:
                    office_file.load_key(password=password)
                    decrypted = io.BytesIO()
                    office_file.decrypt(decrypted)
                    decrypted.seek(0)
                    logger.info(f"文件解密成功: {file_path.name}")
                    return decrypted
                except Exception as e:
                    logger.warning(f"使用密码 {password} 解密失败: {e}")
                    return None
        except Exception as e:
            logger.warning(f"解密文件时出错: {e}")
            return None

    @staticmethod
    def try_common_passwords(file_path: Path) -> Optional[io.BytesIO]:
        """尝试常见密码解密"""
        common_passwords = [
            "VelvetSweatshop",
            "",
            "123456",
            "12345678",
            "password",
            "111111",
            "000000",
            "123123",
        ]
        for pwd in common_passwords:
            result = ExcelParserService.decrypt_file_to_memory(file_path, pwd)
            if result is not None:
                return result
        return None

    @staticmethod
    def _is_header_row(row: tuple) -> bool:
        """判断一行是否像表头（非空单元格数量超过半数）"""
        non_empty = sum(1 for cell in row if cell is not None and str(cell).strip() != "")
        return non_empty > len(row) * 0.5 if row else False

    @staticmethod
    def _find_header_row_index(rows: List[tuple], max_scan: int = 10) -> int:
        """在前面 max_scan 行中找到最像表头的行"""
        best_idx = 0
        best_score = 0

        for i in range(min(max_scan, len(rows))):
            row = rows[i]
            non_empty = sum(1 for cell in row if cell is not None and str(cell).strip() != "")
            if non_empty > best_score and non_empty > 1:
                best_score = non_empty
                best_idx = i

        return best_idx

    @staticmethod
    def _get_data_source(file_path: Path, password: Optional[str] = None):
        """
        获取可用的数据源（文件路径或内存流）
        返回: (source, is_encrypted, password_used)
        """
        is_encrypted = ExcelParserService.is_file_encrypted(file_path)

        if not is_encrypted:
            return file_path, False, None

        if password:
            decrypted = ExcelParserService.decrypt_file_to_memory(file_path, password)
            if decrypted is not None:
                return decrypted, True, password

        decrypted = ExcelParserService.try_common_passwords(file_path)
        if decrypted is not None:
            return decrypted, True, "common_password"

        raise FileEncryptedError(f"文件 {file_path.name} 已加密，请提供正确的密码")

    @staticmethod
    def get_sheet_info(file_path: Path, password: Optional[str] = None) -> List[dict]:
        """获取所有工作表信息"""
        source, is_enc, pwd = ExcelParserService._get_data_source(file_path, password)

        if HAS_CALAMINE and not is_enc:
            try:
                return ExcelParserService._get_sheet_info_calamine(source)
            except Exception as e:
                logger.warning(f"calamine 解析失败，尝试 openpyxl: {e}")

        if HAS_OPENPYXL:
            try:
                return ExcelParserService._get_sheet_info_openpyxl(source)
            except Exception as e:
                logger.warning(f"openpyxl 解析失败，尝试 xlrd: {e}")

        if HAS_XLRD and file_path.suffix.lower() == ".xls":
            try:
                return ExcelParserService._get_sheet_info_xlrd(file_path)
            except Exception as e:
                logger.warning(f"xlrd 解析失败: {e}")

        raise RuntimeError(f"无法解析文件: {file_path.name}，所有解析器均失败")

    @staticmethod
    def _get_sheet_info_calamine(source) -> List[dict]:
        wb = CalamineWorkbook.from_path(str(source) if isinstance(source, Path) else source)
        sheets = []
        for name in wb.sheet_names:
            ws = wb.get_sheet_by_name(name)
            rows = ws.to_python()
            sheets.append({
                "name": name,
                "rowCount": len(rows),
                "columnCount": len(rows[0]) if rows else 0
            })
        return sheets

    @staticmethod
    def _get_sheet_info_openpyxl(source) -> List[dict]:
        wb = openpyxl.load_workbook(source, read_only=True, data_only=True)
        sheets = []
        try:
            for name in wb.sheetnames:
                ws = wb[name]
                sheets.append({
                    "name": name,
                    "rowCount": ws.max_row or 0,
                    "columnCount": ws.max_column or 0
                })
        finally:
            wb.close()
        return sheets

    @staticmethod
    def _get_sheet_info_xlrd(file_path: Path) -> List[dict]:
        wb = xlrd.open_workbook(str(file_path))
        sheets = []
        for i in range(wb.nsheets):
            ws = wb.sheet_by_index(i)
            sheets.append({
                "name": ws.name,
                "rowCount": ws.nrows,
                "columnCount": ws.ncols
            })
        return sheets

    @staticmethod
    def find_main_sheet(file_path: Path, password: Optional[str] = None) -> Optional[str]:
        """找到数据量最大的工作表"""
        sheets = ExcelParserService.get_sheet_info(file_path, password)
        if not sheets:
            return None
        sheets.sort(key=lambda s: s["rowCount"], reverse=True)
        return sheets[0]["name"]

    @staticmethod
    def get_headers(file_path: Path, sheet_name: Optional[str] = None, password: Optional[str] = None) -> List[str]:
        """获取表头"""
        source, is_enc, pwd = ExcelParserService._get_data_source(file_path, password)

        if HAS_CALAMINE and not is_enc:
            try:
                return ExcelParserService._get_headers_calamine(source, sheet_name, file_path)
            except Exception as e:
                logger.warning(f"calamine 解析表头失败，尝试 openpyxl: {e}")

        if HAS_OPENPYXL:
            try:
                return ExcelParserService._get_headers_openpyxl(source, sheet_name)
            except Exception as e:
                logger.warning(f"openpyxl 解析表头失败: {e}")

        raise RuntimeError(f"无法获取表头: {file_path.name}")

    @staticmethod
    def _get_headers_calamine(source, sheet_name: Optional[str], file_path: Path) -> List[str]:
        wb = CalamineWorkbook.from_path(str(source) if isinstance(source, Path) else source)
        if not sheet_name:
            sheet_name = ExcelParserService.find_main_sheet(file_path)
        ws = wb.get_sheet_by_name(sheet_name)
        rows = ws.to_python()

        if not rows:
            return []

        header_idx = ExcelParserService._find_header_row_index(rows)
        header_row = rows[header_idx]

        headers = []
        for i, cell in enumerate(header_row):
            if cell is not None and str(cell).strip() != "":
                headers.append(str(cell).strip())
            else:
                headers.append(f"Column{i+1}")

        return headers

    @staticmethod
    def _get_headers_openpyxl(source, sheet_name: Optional[str]) -> List[str]:
        wb = openpyxl.load_workbook(source, data_only=True)
        try:
            ws = wb[sheet_name] if sheet_name else wb.active
            rows = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= 10:
                    break
                rows.append(tuple(row))

            if not rows:
                return []

            header_idx = ExcelParserService._find_header_row_index(rows)
            header_row = rows[header_idx]

            headers = []
            for i, cell in enumerate(header_row):
                if cell is not None and str(cell).strip() != "":
                    headers.append(str(cell).strip())
                else:
                    headers.append(f"Column{i+1}")
            return headers
        finally:
            wb.close()

    @staticmethod
    def count_rows(file_path: Path, sheet_name: Optional[str] = None, password: Optional[str] = None) -> int:
        """快速统计行数（含表头）"""
        sheets = ExcelParserService.get_sheet_info(file_path, password)
        for s in sheets:
            if s["name"] == sheet_name or (not sheet_name and s["name"] == sheets[0]["name"]):
                return s["rowCount"]
        return 0

    @staticmethod
    def parse_all_sheets(
        file_path: Path,
        password: Optional[str] = None
    ) -> List[dict]:
        """一次性解析所有 sheet，返回 [{name, headers, rows}]，只读取一次文件"""
        source, is_enc, pwd = ExcelParserService._get_data_source(file_path, password)

        if HAS_POLARS and not is_enc:
            try:
                return ExcelParserService._parse_all_polars(source, file_path)
            except Exception as e:
                logger.warning(f"polars 全量解析失败，尝试 calamine: {e}")

        if HAS_CALAMINE and not is_enc:
            try:
                return ExcelParserService._parse_all_calamine(source, file_path)
            except Exception as e:
                logger.warning(f"calamine 全量解析失败，尝试 openpyxl: {e}")

        if HAS_OPENPYXL:
            try:
                return ExcelParserService._parse_all_openpyxl(source, file_path)
            except Exception as e:
                logger.warning(f"openpyxl 全量解析失败: {e}")

        raise RuntimeError(f"无法解析文件: {file_path.name}，所有解析器均失败")

    @staticmethod
    def _parse_all_polars(source, file_path: Path) -> List[dict]:
        if isinstance(source, Path):
            file_path_str = str(source)
        else:
            file_path_str = str(file_path)

        result = pl.read_excel(file_path_str, sheet_id=None)

        if isinstance(result, pl.DataFrame):
            sheets = [("Sheet1", result)]
        else:
            sheets = list(result.items())

        parsed = []
        for name, df in sheets:
            headers = list(df.columns)
            if not headers:
                continue

            all_rows = [tuple(headers)]
            for row in df.iter_rows():
                row_data = []
                for cell in row:
                    if cell is None or (isinstance(cell, float) and cell != cell):
                        row_data.append(None)
                    elif isinstance(cell, float) and cell.is_integer():
                        row_data.append(int(cell))
                    else:
                        row_data.append(cell)
                all_rows.append(tuple(row_data))

            if not all_rows:
                continue

            header_idx = ExcelParserService._find_header_row_index(all_rows)
            parsed_headers = [str(h) if h is not None and str(h).strip() != "" else f"Column{i+1}" for i, h in enumerate(all_rows[header_idx])]
            data_rows = all_rows[header_idx + 1:]

            parsed.append({
                "name": name,
                "headers": parsed_headers,
                "rows": data_rows
            })

        return parsed

    @staticmethod
    def _parse_all_calamine(source, file_path: Path) -> List[dict]:
        wb = CalamineWorkbook.from_path(str(source) if isinstance(source, Path) else source)
        results = []
        for name in wb.sheet_names:
            ws = wb.get_sheet_by_name(name)
            rows = ws.to_python()
            if not rows:
                continue
            header_idx = ExcelParserService._find_header_row_index(rows)
            headers = [str(h) if h is not None and str(h).strip() != "" else f"Column{i+1}" for i, h in enumerate(rows[header_idx])]
            data_rows = [tuple(r) for r in rows[header_idx + 1:]]
            results.append({
                "name": name,
                "headers": headers,
                "rows": data_rows
            })
        return results

    @staticmethod
    def _parse_all_openpyxl(source, file_path: Path) -> List[dict]:
        wb = openpyxl.load_workbook(source, data_only=True)
        try:
            results = []
            for name in wb.sheetnames:
                ws = wb[name]
                all_rows = []
                for row in ws.iter_rows(values_only=True):
                    all_rows.append(tuple(row))
                if not all_rows:
                    continue
                header_idx = ExcelParserService._find_header_row_index(all_rows)
                headers = [str(h) if h is not None and str(h).strip() != "" else f"Column{i+1}" for i, h in enumerate(all_rows[header_idx])]
                data_rows = all_rows[header_idx + 1:]
                results.append({
                    "name": name,
                    "headers": headers,
                    "rows": data_rows
                })
            return results
        finally:
            wb.close()

    @staticmethod
    def parse_file(
        file_path: Path,
        sheet_name: Optional[str] = None,
        chunk_size: int = 1000,
        password: Optional[str] = None
    ) -> Iterator[List[Tuple]]:
        """解析 Excel 文件，按块返回数据（跳过表头行）"""
        source, is_enc, pwd = ExcelParserService._get_data_source(file_path, password)

        if HAS_POLARS and not is_enc:
            try:
                yield from ExcelParserService._parse_polars(source, sheet_name, chunk_size, file_path)
                return
            except Exception as e:
                logger.warning(f"polars 解析失败，尝试 calamine: {e}")

        if HAS_CALAMINE and not is_enc:
            try:
                yield from ExcelParserService._parse_calamine(source, sheet_name, chunk_size, file_path)
                return
            except Exception as e:
                logger.warning(f"calamine 解析失败，尝试 openpyxl: {e}")

        if HAS_OPENPYXL:
            try:
                yield from ExcelParserService._parse_openpyxl(source, sheet_name, chunk_size)
                return
            except Exception as e:
                logger.warning(f"openpyxl 解析失败: {e}")

        raise RuntimeError(f"无法解析文件: {file_path.name}")

    @staticmethod
    def _parse_polars(
        source,
        sheet_name: Optional[str],
        chunk_size: int,
        file_path: Path
    ) -> Iterator[List[Tuple]]:
        if isinstance(source, Path):
            file_path_str = str(source)
        else:
            file_path_str = str(file_path)

        result = pl.read_excel(file_path_str, sheet_id=None)

        if isinstance(result, pl.DataFrame):
            sheets = [("Sheet1", result)]
        else:
            sheets = list(result.items())

        if not sheet_name:
            sheet_name = sheets[0][0]

        for name, df in sheets:
            if name != sheet_name:
                continue

            headers = list(df.columns)
            if not headers:
                return

            all_rows = [tuple(headers)]
            for row in df.iter_rows():
                row_data = []
                for cell in row:
                    if cell is None or (isinstance(cell, float) and cell != cell):
                        row_data.append(None)
                    elif isinstance(cell, float) and cell.is_integer():
                        row_data.append(int(cell))
                    else:
                        row_data.append(cell)
                all_rows.append(tuple(row_data))

            if not all_rows:
                return

            header_idx = ExcelParserService._find_header_row_index(all_rows)
            data_rows = all_rows[header_idx + 1:]

            chunk = []
            for row in data_rows:
                chunk.append(row)
                if len(chunk) >= chunk_size:
                    yield chunk
                    chunk = []

            if chunk:
                yield chunk

            return

        raise RuntimeError(f"工作表 {sheet_name} 不存在")

    @staticmethod
    def _parse_calamine(
        source,
        sheet_name: Optional[str],
        chunk_size: int,
        file_path: Path
    ) -> Iterator[List[Tuple]]:
        wb = CalamineWorkbook.from_path(str(source) if isinstance(source, Path) else source)
        if not sheet_name:
            sheet_name = ExcelParserService.find_main_sheet(file_path)
        ws = wb.get_sheet_by_name(sheet_name)
        rows = ws.to_python()

        if not rows:
            return

        header_idx = ExcelParserService._find_header_row_index(rows)
        data_rows = rows[header_idx + 1:]

        chunk = []
        for row in data_rows:
            chunk.append(tuple(row))
            if len(chunk) >= chunk_size:
                yield chunk
                chunk = []

        if chunk:
            yield chunk

    @staticmethod
    def _parse_openpyxl(
        source,
        sheet_name: Optional[str],
        chunk_size: int
    ) -> Iterator[List[Tuple]]:
        wb = openpyxl.load_workbook(source, data_only=True)
        try:
            ws = wb[sheet_name] if sheet_name else wb.active
            chunk = []
            rows_buffer = []
            header_idx = 0
            found_header = False

            for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                if not found_header:
                    rows_buffer.append(tuple(row))
                    if len(rows_buffer) >= 10:
                        header_idx = ExcelParserService._find_header_row_index(rows_buffer)
                        found_header = True
                        for i in range(header_idx + 1, len(rows_buffer)):
                            chunk.append(rows_buffer[i])
                            if len(chunk) >= chunk_size:
                                yield chunk
                                chunk = []
                    continue

                chunk.append(tuple(row))
                if len(chunk) >= chunk_size:
                    yield chunk
                    chunk = []

            if chunk:
                yield chunk
        finally:
            wb.close()

    @staticmethod
    def parse_csv(file_path: Path, encoding: str = "utf-8") -> Tuple[List[str], List[tuple]]:
        """解析 CSV 文件，返回 (headers, rows)"""
        with open(file_path, "r", encoding=encoding, errors="replace") as f:
            reader = csv.reader(f)
            headers = [str(h) if h else f"Column{i+1}" for i, h in enumerate(next(reader))]
            rows = [tuple(row) for row in reader]
        return headers, rows

    @staticmethod
    def convert_to_str_tuple(row: tuple, expected_length: Optional[int] = None) -> tuple:
        """将行数据转换为字符串元组，确保列数一致"""
        result = []
        for cell in row:
            if cell is None:
                result.append("")
            elif isinstance(cell, float) and cell.is_integer():
                result.append(str(int(cell)))
            elif hasattr(cell, 'strftime'):
                result.append(str(cell))
            else:
                result.append(str(cell))

        if expected_length is not None and len(result) < expected_length:
            result.extend([""] * (expected_length - len(result)))
        elif expected_length is not None and len(result) > expected_length:
            result = result[:expected_length]

        return tuple(result)

    @staticmethod
    def clean_rows(rows: List[tuple]) -> List[tuple]:
        """去除全空行"""
        cleaned = []
        for row in rows:
            has_data = any(cell is not None and str(cell).strip() != "" for cell in row)
            if has_data:
                cleaned.append(row)
        return cleaned

    @staticmethod
    def clean_empty_columns(rows: List[tuple], headers: List[str]) -> Tuple[List[str], List[tuple]]:
        """去除全空列，返回新的 headers 和 rows"""
        if not rows or not headers:
            return headers, rows

        col_count = max(len(headers), max(len(r) for r in rows) if rows else 0)

        empty_cols = set()
        for col_idx in range(col_count):
            is_empty = True
            for row in rows:
                cell = row[col_idx] if col_idx < len(row) else None
                if cell is not None and str(cell).strip() != "":
                    is_empty = False
                    break
            if is_empty:
                empty_cols.add(col_idx)

        if not empty_cols:
            return headers, rows

        new_headers = [h for i, h in enumerate(headers) if i not in empty_cols]

        new_rows = []
        for row in rows:
            new_row = tuple(cell for i, cell in enumerate(row) if i not in empty_cols)
            new_rows.append(new_row)

        return new_headers, new_rows
